#!/usr/bin/env python3
import argparse
import base64
import json
import os
import re
import sqlite3
import subprocess
import threading
import uuid
from io import BytesIO
import urllib.error
import urllib.request
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse, quote
from typing import Optional

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "school_advisor.db"
SOURCES_PATH = ROOT / "config" / "sources.json"
ADMIN_DIR = ROOT / "admin"
SEED_PATH = ROOT / "data" / "seed.json"
SEED_V2_CITY_PATH = ROOT / "data" / "seed_v2_city_shanghai.json"
STRUCTURED_V1_PATH = ROOT / "data" / "curation" / "schools_structured_v1.jsonl"
CANDIDATE_ADMISSION_2025_PATH = ROOT / "data" / "curation" / "admission_2025_candidates.jsonl"
PUBLISH_TASKS: dict[str, dict] = {}
PUBLISH_LOCK = threading.Lock()
ENABLE_1V1_DECISION = False  # 1v1 决策暂时下线，保留接口代码待后续恢复


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    handler.end_headers()
    handler.wfile.write(body)


def _text_response(handler: BaseHTTPRequestHandler, status: int, body: str, content_type: str) -> None:
    data = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def _db_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _db_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS datasets (
              id INTEGER PRIMARY KEY CHECK (id=1),
              sd_json TEXT NOT NULL,
              pr_json TEXT NOT NULL,
              tf_json TEXT NOT NULL,
              dn_json TEXT NOT NULL,
              updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS proposals (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              source TEXT NOT NULL,
              proposal_type TEXT NOT NULL,
              proposal_key TEXT NOT NULL,
              new_value_json TEXT NOT NULL,
              evidence_url TEXT,
              status TEXT NOT NULL DEFAULT 'pending',
              note TEXT,
              created_at TEXT NOT NULL,
              reviewed_at TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_proposals_unique
            ON proposals(source, proposal_type, proposal_key, new_value_json)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS policy_events (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              event_key TEXT NOT NULL UNIQUE,
              event_date TEXT,
              title TEXT NOT NULL,
              source TEXT,
              evidence_url TEXT,
              payload_json TEXT NOT NULL,
              created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS school_evidence (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              school_name TEXT NOT NULL,
              query_text TEXT NOT NULL,
              title TEXT NOT NULL,
              url TEXT NOT NULL,
              snippet TEXT NOT NULL DEFAULT '',
              source_type TEXT NOT NULL DEFAULT 'web',
              published_at TEXT,
              collected_at TEXT NOT NULL,
              status TEXT NOT NULL DEFAULT 'active'
            )
            """
        )
        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_school_evidence_unique
            ON school_evidence(school_name, query_text, url)
            """
        )
        conn.commit()


def _ensure_default_sources() -> None:
    if SOURCES_PATH.exists():
        return
    SOURCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    default = {
        "sources": [
            {
                "id": "sh_municipal_policy",
                "name": "上海市教委官网（政策）",
                "enabled": False,
                "type": "policy_html",
                "url": "https://edu.sh.gov.cn/",
                "keywords": ["幼升小", "义务教育", "招生", "摇号", "入学", "公办", "民办"],
            },
            {
                "id": "district_school_json_sample",
                "name": "区教育局学校 JSON 样例",
                "enabled": False,
                "type": "school_json",
                "url": "https://example.com/schools.json",
                "mapping": {
                    "name": "name",
                    "district": "district",
                    "schoolType": "type",
                    "lotteryLow": "lotteryLow",
                    "lotteryHigh": "lotteryHigh",
                    "recommend": "recommend",
                    "desc": "desc",
                    "lat": "lat",
                    "lng": "lng",
                    "status": "status",
                    "tier": "tier",
                },
            },
        ]
    }
    SOURCES_PATH.write_text(json.dumps(default, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _load_sources() -> list[dict]:
    _ensure_default_sources()
    payload = json.loads(SOURCES_PATH.read_text(encoding="utf-8"))
    sources = payload.get("sources", [])
    if not isinstance(sources, list):
        raise ValueError("sources.json 格式错误：sources 必须是数组")
    return sources


def _normalize_school_name(name: str) -> str:
    return re.sub(r"[（(]\s*小学部\s*[）)]", "小学部", name or "")


def _normalize_payload_names(payload: dict) -> dict:
    sd = payload.get("SD", [])
    pr = payload.get("PR", {})
    tf = payload.get("TF", {})
    dn = payload.get("DN", {})
    updated_at = payload.get("updatedAt")

    name_map: dict[str, str] = {}
    normalized_sd: list = []
    for row in sd:
        if isinstance(row, list) and row:
            old_name = row[0]
            new_name = _normalize_school_name(str(old_name))
            if old_name != new_name:
                name_map[str(old_name)] = new_name
            new_row = list(row)
            new_row[0] = new_name
            normalized_sd.append(new_row)
        else:
            normalized_sd.append(row)

    def _normalize_keyed_map(obj: dict) -> dict:
        out: dict = {}
        for k, v in (obj or {}).items():
            nk = name_map.get(k, _normalize_school_name(k))
            if nk not in out:
                out[nk] = v
        return out

    return {
        "SD": normalized_sd,
        "PR": _normalize_keyed_map(pr),
        "TF": _normalize_keyed_map(tf),
        "DN": dn,
        "updatedAt": updated_at,
    }


def _extract_urls(text: str) -> list[str]:
    raw = str(text or "").strip()
    if not raw:
        return []
    urls = re.findall(r"https?://[^\s<>\u3000]+", raw)
    cleaned = []
    seen = set()
    for u in urls:
        url = str(u).strip().rstrip("；;，,。)")
        if url and url not in seen:
            seen.add(url)
            cleaned.append(url)
    return cleaned


def _xhs_search_url(keyword_or_url: str) -> str:
    raw = str(keyword_or_url or "").strip()
    if not raw:
        return ""
    if raw.startswith("http://") or raw.startswith("https://"):
        return raw
    return "https://www.xiaohongshu.com/search_result?keyword=" + quote(raw)


def _primary_stage_keyword(school_name: str) -> str:
    name = str(school_name or "").strip()
    if not name:
        return ""
    if "小学部" in name:
        return name
    if "小学" in name:
        return name
    return f"{name} 小学"


def _primary_stage_query(school_name: str, *parts: str) -> str:
    base = _primary_stage_keyword(school_name)
    extras = [str(x or "").strip() for x in parts if str(x or "").strip()]
    return " ".join([base] + extras).strip()


def _source_mode(url: str, source_type: str = "", note: str = "") -> str:
    link = str(url or "").strip().lower()
    src_type = str(source_type or "").strip().lower()
    note_text = str(note or "").strip().lower()
    if not link:
        return "lead"
    if any(x in link for x in ("search_result?", "/search?", "baidu.com/s?", "google.com/search", "bing.com/search")):
        return "lead"
    if "xiaohongshu.com" in link and "search_result?" in link:
        return "lead"
    if src_type in ("community-search", "search", "lead"):
        return "lead"
    if "检索" in note_text or "搜索" in note_text:
        return "lead"
    return "evidence"


def _extract_candidate_facts_from_search_text(school_name: str, query_text: str, title: str, snippet: str, url: str = "") -> dict:
    text = " ".join([str(title or "").strip(), str(snippet or "").strip()]).strip()
    out = {
        "schoolName": str(school_name or "").strip(),
        "queryText": str(query_text or "").strip(),
        "title": str(title or "").strip(),
        "url": str(url or "").strip(),
        "stage": "primary" if any(x in str(query_text or "") for x in ("小学", "小学部", "幼升小")) else "unknown",
        "facts": [],
        "warnings": [],
    }
    if not text:
        out["warnings"].append("搜索摘要为空，无法提取候选事实")
        return out

    if any(x in text for x in ("初中", "中学", "高中")) and "小学" not in text and "小学部" not in text:
        out["warnings"].append("摘要可能偏向非小学学段，需人工核验")

    total = re.search(r"(总计|总招生人数|总招生|总录取人数|录取人数|招生总数)[：:\s]*([0-9]{1,4})\s*人", text)
    if total:
        out["facts"].append(
            {
                "field": "admissionTotal2025",
                "value": int(total.group(2)),
                "evidenceText": total.group(0),
                "confidence": "medium",
            }
        )

    max_lottery = re.search(r"(报名人数|报名总数|摇号人数|最大摇号数|报名上限)[：:\s]*([0-9]{1,4})\s*人", text)
    if max_lottery:
        out["facts"].append(
            {
                "field": "maxLottery2025",
                "value": int(max_lottery.group(2)),
                "evidenceText": max_lottery.group(0),
                "confidence": "medium",
            }
        )

    rate_m = re.search(r"(中签率|录取率)[：:\s]*([0-9]{1,3}(?:\.[0-9]+)?)\s*%", text)
    if rate_m:
        out["facts"].append(
            {
                "field": "rate2025",
                "value": float(rate_m.group(2)),
                "evidenceText": rate_m.group(0),
                "confidence": "medium",
            }
        )

    unified_total = re.search(r"统招[：:\s]*([0-9]{1,4})\s*人", text)
    if unified_total:
        out["facts"].append(
            {
                "field": "admission2025",
                "value": int(unified_total.group(1)),
                "evidenceText": unified_total.group(0),
                "confidence": "medium",
            }
        )

    walk_m = re.search(r"统招（走读）[：:\s]*([0-9]{1,4})\s*人", text)
    stay_m = re.search(r"统招（住宿）[：:\s]*([0-9]{1,4})\s*人", text)
    if walk_m and stay_m and not unified_total:
        walk_n = int(walk_m.group(1))
        stay_n = int(stay_m.group(1))
        out["facts"].append(
            {
                "field": "admission2025",
                "value": walk_n + stay_n,
                "evidenceText": f"统招（走读）：{walk_n}人；统招（住宿）：{stay_n}人",
                "confidence": "medium",
            }
        )

    for label, field_key in (
        ("统招（走读）", "admission_boarding_day"),
        ("统招（住宿）", "admission_boarding_live"),
        ("外籍学生", "admission_foreign"),
        ("港澳台学生", "admission_hmt"),
        ("教职工子女", "admission_staff_children"),
    ):
        m = re.search(rf"{re.escape(label)}[：:\s]*([0-9]{{1,4}})\s*人", text)
        if m:
            out["facts"].append(
                {
                    "field": field_key,
                    "value": int(m.group(1)),
                    "evidenceText": m.group(0),
                    "confidence": "medium",
                }
            )

    if "小学" not in str(query_text or "") and "小学部" not in str(query_text or "") and "幼升小" not in str(query_text or ""):
        out["warnings"].append("检索词未限定小学学段，结果可能串到初中或K12")
    if not out["facts"]:
        out["warnings"].append("当前只跑通了招生人数类候选事实提取，其它字段暂未命中")
    return out


def _load_seed_v2_index() -> dict[str, dict]:
    index: dict[str, dict] = {}
    scope_rank = {"city": 1, "district": 2, "school": 3}
    for path in sorted((ROOT / "data").glob("seed_v2_*.json")):
        if not path.is_file():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        scope = str(payload.get("scope") or "city")
        rank = scope_rank.get(scope, 0)
        for school in payload.get("schools", []):
            if not isinstance(school, dict):
                continue
            key = _normalize_school_name(str(school.get("name") or school.get("officialName") or "").strip())
            if not key:
                continue
            prev = index.get(key)
            prev_rank = int(prev.get("__scopeRank", 0)) if isinstance(prev, dict) else 0
            if prev and prev_rank > rank:
                continue
            node = dict(school)
            node["__scope"] = scope
            node["__scopeRank"] = rank
            node["__file"] = path.name
            index[key] = node
    return index


def _load_structured_school_index() -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not STRUCTURED_V1_PATH.exists():
        return out
    for line in STRUCTURED_V1_PATH.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw:
            continue
        try:
            obj = json.loads(raw)
        except Exception:
            continue
        key = _normalize_school_name(str(obj.get("schoolName") or "").strip())
        if key and key not in out:
            out[key] = obj
    return out


def _load_candidate_admission_index() -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not CANDIDATE_ADMISSION_2025_PATH.exists():
        return out
    for line in CANDIDATE_ADMISSION_2025_PATH.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw:
            continue
        try:
            obj = json.loads(raw)
        except Exception:
            continue
        name = _normalize_school_name(str(obj.get("schoolName") or "").strip())
        if not name:
            continue
        existing = out.get(name)
        existing_score = 0
        if isinstance(existing, dict):
            if str(existing.get("factStatus") or "") == "candidate":
                existing_score = 1
            if str(existing.get("confidence") or "") == "high":
                existing_score += 1
        score = 0
        if str(obj.get("factStatus") or "") == "candidate":
            score = 1
        if str(obj.get("confidence") or "") == "high":
            score += 1
        if existing is None or score >= existing_score:
            out[name] = obj
    return out


def _save_sources(sources: list[dict]) -> None:
    SOURCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    SOURCES_PATH.write_text(
        json.dumps({"sources": sources}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _get_bootstrap_payload() -> dict:
    with _db_conn() as conn:
        row = conn.execute(
            "SELECT sd_json, pr_json, tf_json, dn_json, updated_at FROM datasets WHERE id=1"
        ).fetchone()
        if not row:
            return {"SD": [], "PR": {}, "TF": {}, "DN": {}, "updatedAt": None}
        payload = {
            "SD": json.loads(row["sd_json"]),
            "PR": json.loads(row["pr_json"]),
            "TF": json.loads(row["tf_json"]),
            "DN": json.loads(row["dn_json"]),
            "updatedAt": row["updated_at"],
        }
        return _normalize_payload_names(payload)


def _replace_payload(payload: dict) -> None:
    payload = _normalize_payload_names(payload)
    required = ["SD", "PR", "TF", "DN"]
    missing = [k for k in required if k not in payload]
    if missing:
        raise ValueError(f"缺少字段: {', '.join(missing)}")
    if not isinstance(payload["SD"], list):
        raise ValueError("SD 必须是数组")
    if not isinstance(payload["PR"], dict):
        raise ValueError("PR 必须是对象")
    if not isinstance(payload["TF"], dict):
        raise ValueError("TF 必须是对象")
    if not isinstance(payload["DN"], dict):
        raise ValueError("DN 必须是对象")

    with _db_conn() as conn:
        conn.execute(
            """
            INSERT INTO datasets (id, sd_json, pr_json, tf_json, dn_json, updated_at)
            VALUES (1, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
              sd_json=excluded.sd_json,
              pr_json=excluded.pr_json,
              tf_json=excluded.tf_json,
              dn_json=excluded.dn_json,
              updated_at=excluded.updated_at
            """,
            (
                json.dumps(payload["SD"], ensure_ascii=False),
                json.dumps(payload["PR"], ensure_ascii=False),
                json.dumps(payload["TF"], ensure_ascii=False),
                json.dumps(payload["DN"], ensure_ascii=False),
                _now(),
            ),
        )
        conn.commit()


def _write_seed_json(payload: dict) -> None:
    SEED_PATH.parent.mkdir(parents=True, exist_ok=True)
    SEED_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _run_git(args: list[str], extra_env: Optional[dict] = None, timeout_sec: int = 120) -> tuple[int, str]:
    env = None
    if extra_env:
        env = dict(**__import__("os").environ)
        env.update(extra_env)
    try:
        p = subprocess.run(
            args,
            cwd=str(ROOT),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            timeout=timeout_sec,
        )
        return p.returncode, p.stdout
    except subprocess.TimeoutExpired as e:
        out = (e.stdout or "") + "\n" + (e.stderr or "")
        return 124, f"命令超时（>{timeout_sec}s）\n{out}".strip()


def _publish_online(message: str, remote: str, refspec: str, git_ssh_command: str = "") -> dict:
    payload = _get_bootstrap_payload()
    _write_seed_json(payload)

    logs = []
    env = {}
    if git_ssh_command.strip():
        env["GIT_SSH_COMMAND"] = git_ssh_command.strip()

    code, out = _run_git(["git", "add", "data/seed.json"], extra_env=env)
    logs.append({"cmd": "git add data/seed.json", "code": code, "output": out.strip()})
    if code != 0:
        return {"ok": False, "step": "git_add", "logs": logs}

    code, out = _run_git(["git", "diff", "--cached", "--quiet"], extra_env=env)
    has_changes = code != 0
    logs.append({"cmd": "git diff --cached --quiet", "code": code, "output": out.strip()})

    committed = False
    if has_changes:
        code, out = _run_git(["git", "commit", "-m", message], extra_env=env)
        logs.append({"cmd": f'git commit -m "{message}"', "code": code, "output": out.strip()})
        if code != 0:
            return {"ok": False, "step": "git_commit", "logs": logs}
        committed = True

    code, out = _run_git(["git", "push", remote, refspec], extra_env=env)
    logs.append({"cmd": f"git push {remote} {refspec}", "code": code, "output": out.strip()})
    if code != 0:
        return {"ok": False, "step": "git_push", "logs": logs, "committed": committed}

    return {"ok": True, "logs": logs, "committed": committed}


def _publish_to_cloudflare_api(url: str, token: str, payload: dict, timeout_sec: int = 30) -> tuple[int, str]:
    if not url or not token:
        return 400, "cloudflare url/token 不能为空"
    try:
        u = urlparse(url)
        origin = f"{u.scheme}://{u.netloc}" if u.scheme and u.netloc else "https://school-advisor.pages.dev"
    except Exception:
        origin = "https://school-advisor.pages.dev"
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36",
            "Origin": origin,
            "Authorization": f"Bearer {token}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            return resp.getcode(), resp.read().decode("utf-8", errors="ignore")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="ignore")
    except Exception as e:
        return 599, str(e)


def _new_publish_task(message: str, remote: str, refspec: str, mode: str = "git", cloudflare_url: str = "") -> dict:
    task_id = str(uuid.uuid4())
    now = _now()
    steps = [{"key": "export_seed", "label": "导出数据库到 seed.json", "status": "pending", "detail": ""}]
    if mode == "api":
        steps.append({"key": "api_publish", "label": "推送到 Cloudflare API", "status": "pending", "detail": ""})
    else:
        steps.extend(
            [
                {"key": "git_add", "label": "Git 暂存文件", "status": "pending", "detail": ""},
                {"key": "git_check", "label": "检查是否有变更", "status": "pending", "detail": ""},
                {"key": "git_commit", "label": "创建提交", "status": "pending", "detail": ""},
                {"key": "git_push", "label": "推送到线上仓库", "status": "pending", "detail": ""},
            ]
        )
    task = {
        "taskId": task_id,
        "status": "pending",
        "message": message,
        "mode": mode,
        "remote": remote,
        "refspec": refspec,
        "cloudflareUrl": cloudflare_url,
        "createdAt": now,
        "updatedAt": now,
        "error": "",
        "logs": [],
        "steps": steps,
    }
    with PUBLISH_LOCK:
        PUBLISH_TASKS[task_id] = task
    return task


def _snapshot_publish_task(task_id: str) -> Optional[dict]:
    with PUBLISH_LOCK:
        task = PUBLISH_TASKS.get(task_id)
        if not task:
            return None
        return json.loads(json.dumps(task))


def _update_publish_task(task_id: str, **kwargs) -> None:
    with PUBLISH_LOCK:
        task = PUBLISH_TASKS.get(task_id)
        if not task:
            return
        task.update(kwargs)
        task["updatedAt"] = _now()


def _set_publish_step(task_id: str, key: str, status: str, detail: str = "") -> None:
    with PUBLISH_LOCK:
        task = PUBLISH_TASKS.get(task_id)
        if not task:
            return
        for step in task.get("steps", []):
            if step.get("key") == key:
                step["status"] = status
                if detail:
                    step["detail"] = detail
                break
        task["updatedAt"] = _now()


def _append_publish_log(task_id: str, cmd: str, code: int, output: str) -> None:
    with PUBLISH_LOCK:
        task = PUBLISH_TASKS.get(task_id)
        if not task:
            return
        task["logs"].append({"time": _now(), "cmd": cmd, "code": code, "output": (output or "").strip()})
        task["updatedAt"] = _now()


def _run_publish_task(
    task_id: str,
    message: str,
    remote: str,
    refspec: str,
    git_ssh_command: str = "",
    mode: str = "git",
    cloudflare_url: str = "",
    publish_token: str = "",
) -> None:
    _update_publish_task(task_id, status="running")
    env = {}
    if git_ssh_command.strip():
        env["GIT_SSH_COMMAND"] = git_ssh_command.strip()

    try:
        _set_publish_step(task_id, "export_seed", "running")
        payload = _get_bootstrap_payload()
        _write_seed_json(payload)
        _set_publish_step(task_id, "export_seed", "success")

        if mode == "api":
            _set_publish_step(task_id, "api_publish", "running")
            code, out = _publish_to_cloudflare_api(cloudflare_url, publish_token, payload, timeout_sec=35)
            _append_publish_log(task_id, f"POST {cloudflare_url}", code, out)
            if code < 200 or code >= 300:
                _set_publish_step(task_id, "api_publish", "failed", "Cloudflare API 推送失败")
                _update_publish_task(task_id, status="failed", error=(out or "").strip())
                return
            _set_publish_step(task_id, "api_publish", "success")
            _update_publish_task(task_id, status="success")
            return

        _set_publish_step(task_id, "git_add", "running")
        code, out = _run_git(["git", "add", "data/seed.json"], extra_env=env)
        _append_publish_log(task_id, "git add data/seed.json", code, out)
        if code != 0:
            _set_publish_step(task_id, "git_add", "failed", "git add 失败")
            _update_publish_task(task_id, status="failed", error=(out or "").strip())
            return
        _set_publish_step(task_id, "git_add", "success")

        _set_publish_step(task_id, "git_check", "running")
        code, out = _run_git(["git", "diff", "--cached", "--quiet"], extra_env=env)
        _append_publish_log(task_id, "git diff --cached --quiet", code, out)
        has_changes = code != 0
        _set_publish_step(task_id, "git_check", "success", "有变更" if has_changes else "无文件变更，继续执行推送")

        if has_changes:
            _set_publish_step(task_id, "git_commit", "running")
            code, out = _run_git(["git", "commit", "-m", message], extra_env=env)
            _append_publish_log(task_id, f'git commit -m "{message}"', code, out)
            if code != 0:
                _set_publish_step(task_id, "git_commit", "failed", "git commit 失败")
                _update_publish_task(task_id, status="failed", error=(out or "").strip())
                return
            _set_publish_step(task_id, "git_commit", "success")
        else:
            _set_publish_step(task_id, "git_commit", "success", "无需提交")

        _set_publish_step(task_id, "git_push", "running")
        code, out = _run_git(["git", "push", remote, refspec], extra_env=env, timeout_sec=90)
        _append_publish_log(task_id, f"git push {remote} {refspec}", code, out)
        if code != 0:
            _set_publish_step(task_id, "git_push", "failed", "git push 失败")
            _update_publish_task(task_id, status="failed", error=(out or "").strip())
            return
        _set_publish_step(task_id, "git_push", "success")
        _update_publish_task(task_id, status="success")
    except Exception as e:
        _update_publish_task(task_id, status="failed", error=str(e))


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v, default=None):
    if v is None:
        return default
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


def _normalize_type(v) -> str:
    s = str(v or "").strip().lower()
    if s in ("pub", "public", "公办"):
        return "pub"
    if s in ("pri", "private", "民办"):
        return "pri"
    return "pub"


def _normalize_tier(v) -> str:
    s = str(v or "").strip().upper().replace(" ", "")
    if s in ("T1", "T2", "T3"):
        return s
    if s in ("1", "2", "3"):
        return f"T{s}"
    return "T3"


def _normalize_status(v) -> str:
    s = str(v or "").strip().lower()
    if s in ("hot", "超额", "超额摇号"):
        return "hot"
    if s in ("cool", "全录", "全部录取"):
        return "cool"
    return "normal"


def _build_school_items(payload: dict) -> list[dict]:
    items = []
    sd = payload.get("SD", [])
    pr = payload.get("PR", {})
    tf = payload.get("TF", {})
    candidate_index = _load_candidate_admission_index()
    for row in sd:
        if not isinstance(row, list) or not row:
            continue
        name = row[0]
        info = pr.get(name) if isinstance(pr, dict) else None
        fee = tf.get(name) if isinstance(tf, dict) else None
        school_type = row[2] if len(row) > 2 else ""
        official_admission = row[11] if len(row) > 11 else None
        source_url = row[13] if len(row) > 13 else ""
        candidate = _safe_obj(candidate_index.get(_normalize_school_name(str(name))))
        candidate_admission = _to_int(candidate.get("admission2025"), default=None)
        candidate_total_admission = _to_int(candidate.get("admissionTotal2025"), default=None)
        candidate_max_lottery = _to_int(candidate.get("maxLottery2025"), default=None)
        if school_type == "pub" and official_admission is None and candidate_admission is None:
            admission_status = "not_applicable"
            display_admission = None
        elif official_admission is not None and source_url:
            admission_status = "official"
            display_admission = official_admission
        elif candidate_admission is not None:
            admission_status = "candidate"
            display_admission = candidate_admission
        elif official_admission is not None:
            admission_status = "candidate"
            display_admission = official_admission
        else:
            admission_status = "missing"
            display_admission = None
        items.append(
            {
                "name": name,
                "district": row[1] if len(row) > 1 else "",
                "type": school_type,
                "lotteryLow": row[3] if len(row) > 3 else None,
                "lotteryHigh": row[4] if len(row) > 4 else None,
                "recommend": row[5] if len(row) > 5 else None,
                "desc": row[6] if len(row) > 6 else "",
                "lat": row[7] if len(row) > 7 else None,
                "lng": row[8] if len(row) > 8 else None,
                "status": row[9] if len(row) > 9 else "normal",
                "tier": row[10] if len(row) > 10 else "T3",
                "admission2025": display_admission,
                "admission2025Raw": official_admission,
                "admission2025Candidate": candidate_admission,
                "admissionTotal2025Candidate": candidate_total_admission,
                "admission2025Status": admission_status,
                "admission2025EvidenceUrl": str(candidate.get("evidenceUrl") or source_url or "").strip(),
                "admission2025EvidenceText": str(candidate.get("evidenceText") or "").strip(),
                "admission2025Confidence": str(candidate.get("confidence") or "").strip(),
                "maxLottery2025": row[12] if len(row) > 12 and row[12] is not None else candidate_max_lottery,
                "maxLottery2025Raw": row[12] if len(row) > 12 else None,
                "maxLottery2025Candidate": candidate_max_lottery,
                "sourceUrl": source_url,
                "profile": info or {},
                "tuition": fee or {},
            }
        )
    return items


def _school_field_sections(item: dict) -> list[dict]:
    it = _safe_obj(item)
    p = _safe_obj(it.get("profile"))
    source_url = _to_str(it.get("sourceUrl")).strip()
    has_coord = it.get("lat") is not None and it.get("lng") is not None
    lottery_text = f"{it.get('lotteryLow') if it.get('lotteryLow') is not None else '-'} ~ {it.get('lotteryHigh') if it.get('lotteryHigh') is not None else '-'}"
    hw_val = p.get("hw") if p.get("hw") is not None else "-"
    stress_val = p.get("stress") if p.get("stress") is not None else "-"
    school_name = str(it.get("name") or "").strip()
    school_query_name = _primary_stage_keyword(school_name)
    tuition = _safe_obj(it.get("tuition"))
    if tuition:
        tuition_text = f"{tuition.get('term') if tuition.get('term') is not None else '-'} /学期；{tuition.get('note') or '-'}"
    else:
        tuition_text = "-"

    def _xhs_search(keyword: str) -> str:
        return "https://www.xiaohongshu.com/search_result?keyword=" + quote(keyword)

    def _web_search(keyword: str) -> str:
        return "https://www.baidu.com/s?wd=" + quote(keyword)

    def src(current: str, target: str, origin: str, method: str, url: str = "", links: Optional[list[dict]] = None) -> dict:
        link_list = _safe_list(links)
        if url:
            link_list = [{"label": "主来源", "url": url}] + link_list
        return {
            "currentLevel": _safe_source_level(current),
            "targetLevel": _safe_source_level(target),
            "origin": origin,
            "method": method,
            "url": url or None,
            "links": [
                {"label": str(x.get("label") or "来源"), "url": str(x.get("url") or "")}
                for x in link_list
                if isinstance(x, dict) and str(x.get("url") or "").strip()
            ],
        }

    return [
        {
            "category": "identity",
            "title": "第一类 学校身份及基础信息",
            "fields": [
                {
                    "key": "schoolName",
                    "label": "学校",
                    "value": it.get("name") or "-",
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "教育局/学校官网名称口径",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "招生简章"))}],
                    ),
                },
                {
                    "key": "district",
                    "label": "区域",
                    "value": _infer_district_label(_to_str(it.get("district"))),
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "教育局区划字段",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(f"{school_name} 所在区 教育局")}],
                    ),
                },
                {
                    "key": "schoolType",
                    "label": "类型",
                    "value": "公办" if _to_str(it.get("type")) == "pub" else "民办" if _to_str(it.get("type")) == "pri" else "-",
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "招生简章/教育局办学性质",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "办学性质 公办 民办"))}],
                    ),
                },
                {
                    "key": "tier",
                    "label": "梯队",
                    "value": it.get("tier") or "-",
                    "source": src(
                        "community",
                        "verified",
                        "小红书社区",
                        "检索“XX区小学+梯队”，聚合多篇内容统计",
                        "",
                        [{"label": "社区检索", "url": _xhs_search(f"{_infer_district_label(_to_str(it.get('district')))} 小学 梯队")}],
                    ),
                },
                {
                    "key": "coord",
                    "label": "坐标",
                    "value": f"{it.get('lat') if it.get('lat') is not None else '-'}, {it.get('lng') if it.get('lng') is not None else '-'}",
                    "source": src(
                        "official" if has_coord else "ai-draft",
                        "official",
                        "官方地址/地理编码",
                        "基于学校官方地址地理编码",
                        source_url,
                        [{"label": "地图检索", "url": _web_search(_primary_stage_query(school_name, "地址"))}],
                    ),
                },
            ],
        },
        {
            "category": "admission",
            "title": "第二类 招生数据信息",
            "fields": [
                {
                    "key": "lotteryRange",
                    "label": "中签率",
                    "value": lottery_text,
                    "source": src(
                        "community",
                        "verified",
                        "小红书社区",
                        "按 2025录取数 / 2025最大摇号数 计算",
                        "",
                        [{"label": "社区检索", "url": _xhs_search(_primary_stage_query(school_name, "2025", "幼升小", "摇号", "中签率"))}],
                    ),
                },
                {
                    "key": "admission2025",
                    "label": "2025录取数",
                    "value": it.get("admission2025") if it.get("admission2025") is not None else "-",
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "教育局/学校发布口径",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "2025", "幼升小", "中签率"))}],
                    ),
                },
                {
                    "key": "maxLottery2025",
                    "label": "2025最大摇号数",
                    "value": it.get("maxLottery2025") if it.get("maxLottery2025") is not None else "-",
                    "source": src(
                        "community",
                        "verified",
                        "小红书图表识别",
                        "家长图表OCR识别后人工复核",
                        "",
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "2025", "幼升小", "中签率"))}],
                    ),
                },
            ],
        },
        {
            "category": "profile",
            "title": "第三类 详细介绍及口碑信息",
            "fields": [
                {
                    "key": "desc",
                    "label": "简介",
                    "value": it.get("desc") or "-",
                    "source": src(
                        "ai-draft",
                        "verified",
                        "AI综合生成",
                        "多源信息聚合后的中性介绍",
                        "",
                        [{"label": "支撑检索", "url": _web_search(_primary_stage_query(school_name, "学校介绍"))}],
                    ),
                },
                {
                    "key": "philosophy",
                    "label": "教育理念",
                    "value": p.get("slogan") or "-",
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "学校官网办学理念/课程理念",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "办学理念"))}],
                    ),
                },
                {
                    "key": "hwStress",
                    "label": "作业/压力",
                    "value": f"{hw_val} / {stress_val}",
                    "source": src(
                        "community",
                        "verified",
                        "小红书社区",
                        "检索“小学/小学部+作业+压力”，取高相关内容聚合统计",
                        "",
                        [{"label": "社区检索", "url": _xhs_search(_primary_stage_query(school_name, "作业", "压力"))}],
                    ),
                },
                {
                    "key": "path",
                    "label": "升学路径",
                    "value": _safe_list(p.get("path")),
                    "source": src(
                        "official",
                        "official",
                        "官方口径",
                        "集团校/对口/招生说明",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "对口", "初中"))}],
                    ),
                },
                {
                    "key": "pros",
                    "label": "优点",
                    "value": _safe_list(p.get("pros")),
                    "source": src(
                        "ai-draft",
                        "verified",
                        "AI综合生成",
                        "基于口碑与家长关注维度生成优势摘要",
                        "",
                        [{"label": "支撑检索", "url": _xhs_search(_primary_stage_query(school_name, "开放日", "体验"))}],
                    ),
                },
                {
                    "key": "cons",
                    "label": "注意点",
                    "value": _safe_list(p.get("cons")),
                    "source": src(
                        "community",
                        "verified",
                        "社区口碑/咨询复盘",
                        "小红书与家长反馈聚合后整理",
                        "",
                        [{"label": "社区检索", "url": _xhs_search(_primary_stage_query(school_name, "家长", "反馈", "坑点"))}],
                    ),
                },
                {
                    "key": "tuition",
                    "label": "学费信息",
                    "value": tuition_text,
                    "source": src(
                        "official" if tuition else "ai-draft",
                        "official",
                        "学费标准",
                        "招生简章/收费公示",
                        source_url,
                        [{"label": "检索入口", "url": _web_search(_primary_stage_query(school_name, "学费", "收费标准"))}],
                    ),
                },
            ],
        },
    ]


def _safe_obj(v):
    return v if isinstance(v, dict) else {}


def _safe_list(v):
    return v if isinstance(v, list) else []


def _to_str(v) -> str:
    return v if isinstance(v, str) else ""


def _trim_chat_messages(messages) -> list[dict]:
    out = []
    for m in _safe_list(messages):
        if not isinstance(m, dict):
            continue
        role = "assistant" if str(m.get("role") or "").strip() == "assistant" else "user"
        content = str(m.get("content") or "").strip()
        if not content:
            continue
        out.append({"role": role, "content": content})
    return out[-12:]


def _infer_district_label(code: str) -> str:
    return {
        "pudong": "浦东新区",
        "minhang": "闵行区",
        "xuhui": "徐汇区",
        "changning": "长宁区",
        "jingan": "静安区",
        "huangpu": "黄浦区",
        "putuo": "普陀区",
        "yangpu": "杨浦区",
        "hongkou": "虹口区",
        "jiading": "嘉定区",
        "qingpu": "青浦区",
    }.get(code or "", code or "未提供")


def _profile_brief(profile: dict) -> str:
    p = _safe_obj(profile)
    focus = [str(x) for x in _safe_list(p.get("focus")) if str(x).strip()]
    return (
        f"户籍={str(p.get('household') or '未提供')}；"
        f"区域={_infer_district_label(str(p.get('district') or ''))}；"
        f"意向={str(p.get('intentType') or '未提供')}；"
        f"关注={'/'.join(focus) if focus else '未提供'}"
    )


def _district_school_hints(payload: dict, district_code: str) -> list[dict]:
    sd = _safe_list(_safe_obj(payload).get("SD"))
    if not district_code:
        return []
    out = []
    for r in sd:
        if not (isinstance(r, list) and len(r) >= 2 and str(r[1]) == district_code):
            continue
        admission = _to_int(r[11], default=None) if len(r) > 11 else None
        max_lottery = _to_int(r[12], default=None) if len(r) > 12 else None
        if admission is not None and max_lottery and max_lottery > 0:
            lottery = f"{round(admission * 1000 / max_lottery) / 10}%"
        elif len(r) > 3 and isinstance(r[3], (int, float)):
            lottery = f"{r[3]}%"
        else:
            lottery = "站内暂无中签率"
        out.append(
            {
                "name": str(r[0]) if len(r) > 0 else "",
                "type": "公办" if len(r) > 2 and r[2] == "pub" else "民办" if len(r) > 2 and r[2] == "pri" else "未知",
                "tier": str(r[10]) if len(r) > 10 and r[10] else "未知",
                "admission": admission,
                "maxLottery": max_lottery,
                "lottery": lottery,
            }
        )
    return out[:10]


def _evidence_query_templates(school_name: str) -> list[str]:
    q = _primary_stage_keyword(school_name)
    if not q:
        return []
    return [
        f"{q} 2025 幼升小 中签率",
        f"{q} 2025 幼升小 录取数",
        f"{q} 2025 小学 招生计划",
        f"{q} 开放日体验",
        f"{q} 作业 压力",
        f"{q} 办学理念",
    ]


def _search_evidence(query_text: str, top: int = 3) -> list[dict]:
    endpoint = (os.getenv("EVIDENCE_SEARCH_ENDPOINT") or "").strip()
    if not endpoint:
        return []
    token = (os.getenv("EVIDENCE_SEARCH_API_KEY") or "").strip()
    url = endpoint.rstrip("/") + "?" + urlencode({"q": query_text, "top": max(1, min(8, int(top or 3)))})
    headers = {"User-Agent": "SchoolAdvisorEvidenceBot/1.0"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=20) as resp:
        raw = resp.read().decode("utf-8", errors="ignore")
    data = json.loads(raw or "{}")
    rows = data.get("results")
    if not isinstance(rows, list):
        rows = data.get("items")
    if not isinstance(rows, list):
        rows = data.get("data")
    if not isinstance(rows, list):
        return []
    out = []
    for x in rows[:top]:
        if not isinstance(x, dict):
            continue
        title = str(x.get("title") or x.get("name") or "").strip()
        link = str(x.get("url") or x.get("link") or "").strip()
        if not title or not link:
            continue
        out.append(
            {
                "title": title[:180],
                "url": link[:800],
                "snippet": str(x.get("snippet") or x.get("summary") or x.get("description") or "").strip()[:400],
                "sourceType": str(x.get("source_type") or x.get("source") or "web").strip()[:40] or "web",
                "publishedAt": str(x.get("published_at") or x.get("date") or x.get("publish_time") or "").strip()[:40] or None,
            }
        )
    return out


def _upsert_school_evidence_rows(school_name: str, query_text: str, rows: list[dict]) -> int:
    if not rows:
        return 0
    now = _now()
    changed = 0
    with _db_conn() as conn:
        for r in rows:
            try:
                conn.execute(
                    """
                    INSERT INTO school_evidence(
                      school_name, query_text, title, url, snippet, source_type, published_at, collected_at, status
                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, 'active')
                    ON CONFLICT(school_name, query_text, url)
                    DO UPDATE SET
                      title=excluded.title,
                      snippet=excluded.snippet,
                      source_type=excluded.source_type,
                      published_at=excluded.published_at,
                      collected_at=excluded.collected_at,
                      status='active'
                    """,
                    (
                        school_name,
                        query_text,
                        str(r.get("title") or ""),
                        str(r.get("url") or ""),
                        str(r.get("snippet") or ""),
                        str(r.get("sourceType") or "web"),
                        r.get("publishedAt"),
                        now,
                    ),
                )
                changed += 1
            except Exception:
                pass
        conn.commit()
    return changed


def _list_school_evidence(school_name: str, limit: int = 50) -> list[dict]:
    lim = max(1, min(200, int(limit or 50)))
    with _db_conn() as conn:
        rows = conn.execute(
            """
            SELECT id, school_name, query_text, title, url, snippet, source_type, published_at, collected_at, status
            FROM school_evidence
            WHERE school_name = ? AND status = 'active'
            ORDER BY collected_at DESC, id DESC
            LIMIT ?
            """,
            (school_name, lim),
        ).fetchall()
    return [
        {
            "id": r["id"],
            "schoolName": r["school_name"],
            "queryText": r["query_text"],
            "title": r["title"],
            "url": r["url"],
            "snippet": r["snippet"],
            "sourceType": r["source_type"],
            "publishedAt": r["published_at"],
            "collectedAt": r["collected_at"],
            "status": r["status"],
        }
        for r in rows
    ]


def _collect_school_evidence(school_name: str, extra_queries: Optional[list[str]] = None, top: int = 3) -> dict:
    if not school_name:
        return {"schoolName": school_name, "queries": 0, "fetched": 0, "saved": 0, "errors": []}
    templates = _evidence_query_templates(school_name)
    for q in _safe_list(extra_queries):
        qq = str(q or "").strip()
        if qq:
            templates.append(qq)
    # de-dup by exact query
    seen = set()
    queries = []
    for q in templates:
        if q in seen:
            continue
        seen.add(q)
        queries.append(q)
    fetched_total = 0
    saved_total = 0
    errors = []
    for q in queries:
        try:
            rows = _search_evidence(q, top=top)
            fetched_total += len(rows)
            saved_total += _upsert_school_evidence_rows(school_name, q, rows)
        except Exception as e:
            errors.append({"query": q, "error": str(e)})
    return {
        "schoolName": school_name,
        "queries": len(queries),
        "fetched": fetched_total,
        "saved": saved_total,
        "errors": errors,
    }


_SOURCE_LEVELS = ("official", "verified", "community", "ai-draft")


def _safe_source_level(v: str) -> str:
    s = str(v or "").strip().lower()
    if s in _SOURCE_LEVELS:
        return s
    return "ai-draft"


def _guess_level_from_url(url: str) -> str:
    u = str(url or "").strip().lower()
    if not u:
        return "ai-draft"
    if any(x in u for x in ("edu.sh.gov.cn", ".gov.cn", "zsb.", "zs", "admission")):
        return "official"
    if any(x in u for x in ("xiaohongshu.com", "xhslink.com", "zhihu.com", "weibo.com")):
        return "community"
    return "verified"


def _collect_school_sources(item: dict) -> dict:
    school_name = _normalize_school_name(str(_safe_obj(item).get("name") or "").strip())
    profile = _safe_obj(_safe_obj(item).get("profile"))
    seed_v2_index = _load_seed_v2_index()
    structured_index = _load_structured_school_index()
    v2 = _safe_obj(seed_v2_index.get(school_name))
    structured = _safe_obj(structured_index.get(school_name))

    school_sources = []
    field_sources = []
    evidence_sources = []
    school_seen = set()
    field_seen = set()
    evidence_seen = set()

    def push_school_source(title: str, url: str, level: str, source_type: str, origin: str, note: str = "") -> None:
        link = str(url or "").strip()
        if not link:
            return
        key = (link, str(title or "").strip())
        if key in school_seen:
            return
        school_seen.add(key)
        school_sources.append(
            {
                "title": str(title or "来源").strip(),
                "url": link,
                "level": _safe_source_level(level),
                "sourceType": str(source_type or "web").strip() or "web",
                "mode": _source_mode(link, source_type=source_type, note=note),
                "origin": str(origin or "").strip() or "-",
                "note": str(note or "").strip(),
            }
        )

    def push_field_source(
        category: str,
        field_key: str,
        field_label: str,
        value,
        url: str,
        current_level: str,
        target_level: str,
        origin: str,
        method: str,
        label: str = "",
    ) -> None:
        link = str(url or "").strip()
        if not link:
            return
        key = (category, field_key, link)
        if key in field_seen:
            return
        field_seen.add(key)
        field_sources.append(
            {
                "category": str(category or "").strip() or "unknown",
                "fieldKey": str(field_key or "").strip() or "unknown",
                "fieldLabel": str(field_label or "字段").strip(),
                "value": value,
                "url": link,
                "linkLabel": str(label or "来源").strip(),
                "mode": _source_mode(link, source_type=label, note=method),
                "currentLevel": _safe_source_level(current_level),
                "targetLevel": _safe_source_level(target_level),
                "origin": str(origin or "").strip() or "-",
                "method": str(method or "").strip() or "-",
            }
        )

    source_url = str(_safe_obj(item).get("sourceUrl") or "").strip()
    if source_url:
        push_school_source("主数据 sourceUrl", source_url, _guess_level_from_url(source_url), "bootstrap", "seed.json / SD[13]")

    pr_xhs = str(profile.get("xhs") or "").strip()
    if pr_xhs:
        push_school_source("PR 小红书检索", _xhs_search_url(pr_xhs), "community", "community-search", "seed.json / PR.xhs", pr_xhs)

    admission = _safe_obj(v2.get("admission"))
    admission_url = str(admission.get("admissionUrl") or "").strip()
    if admission_url:
        push_school_source(
            "招生来源",
            admission_url,
            str(admission.get("admissionSource") or _guess_level_from_url(admission_url)),
            "admission",
            f"{v2.get('__file') or 'seed_v2'} / admission.admissionUrl",
        )

    links = _safe_obj(v2.get("links"))
    xhs_link = str(links.get("xhs") or "").strip()
    if xhs_link:
        push_school_source("seed_v2 小红书检索", _xhs_search_url(xhs_link), "community", "community-search", f"{v2.get('__file') or 'seed_v2'} / links.xhs", xhs_link)

    profile_v2 = _safe_obj(v2.get("profile"))
    source_note = str(profile_v2.get("sourceNote") or "").strip()
    for idx, url in enumerate(_extract_urls(source_note), start=1):
        guessed_level = _guess_level_from_url(url)
        push_school_source(
            f"profile.sourceNote 来源 {idx}",
            url,
            guessed_level if guessed_level != "ai-draft" else str(profile_v2.get("sourceLevel") or guessed_level),
            "profile-note",
            f"{v2.get('__file') or 'seed_v2'} / profile.sourceNote",
        )

    for cat in _safe_list(structured.get("categories")):
        category = str(_safe_obj(cat).get("category") or "").strip()
        for field in _safe_list(_safe_obj(cat).get("fields")):
            field_obj = _safe_obj(field)
            field_key = str(field_obj.get("key") or "").strip()
            field_label = str(field_obj.get("label") or field_key or "字段").strip()
            field_value = field_obj.get("value")
            current_level = str(field_obj.get("currentLevel") or "")
            target_level = str(field_obj.get("targetLevel") or current_level or "verified")
            origin = str(field_obj.get("origin") or "")
            method = str(field_obj.get("method") or "")
            for link_obj in _safe_list(field_obj.get("links")):
                link = _safe_obj(link_obj)
                push_field_source(
                    category,
                    field_key,
                    field_label,
                    field_value,
                    str(link.get("url") or "").strip(),
                    current_level,
                    target_level,
                    origin,
                    method,
                    str(link.get("label") or "来源"),
                )

    for ev in _list_school_evidence(school_name, limit=60):
        url = str(_safe_obj(ev).get("url") or "").strip()
        if not url or url in evidence_seen:
            continue
        evidence_seen.add(url)
        evidence_sources.append(
            {
                "title": str(ev.get("title") or "").strip() or "证据",
                "url": url,
                "queryText": str(ev.get("queryText") or "").strip(),
                "sourceType": str(ev.get("sourceType") or "web").strip() or "web",
                "mode": "evidence",
                "publishedAt": ev.get("publishedAt"),
                "collectedAt": ev.get("collectedAt"),
                "snippet": str(ev.get("snippet") or "").strip(),
            }
        )

    return {
        "summary": {
            "schoolSourceCount": len(school_sources),
            "fieldSourceCount": len(field_sources),
            "evidenceCount": len(evidence_sources),
        },
        "schoolSources": school_sources,
        "fieldSources": field_sources,
        "evidenceSources": evidence_sources,
    }


def _infer_school_provenance(row: list, pr_node: dict) -> dict:
    source_url = ""
    if isinstance(row, list) and len(row) > 13:
        source_url = str(row[13] or "").strip()
    source_level = _guess_level_from_url(source_url)
    pr_obj = _safe_obj(pr_node)
    explicit_level = _safe_source_level(pr_obj.get("sourceLevel") or "")
    if explicit_level != "ai-draft":
        reputation_level = explicit_level
    elif _safe_list(pr_obj.get("xhsSignals")):
        reputation_level = "community"
    elif source_level in ("official", "verified"):
        reputation_level = "verified"
    else:
        reputation_level = "ai-draft"

    # 理念字段默认沿用口碑来源级别；若未来有 philosophySourceLevel 则优先使用
    philosophy_level = _safe_source_level(pr_obj.get("philosophySourceLevel") or "")
    if philosophy_level == "ai-draft":
        philosophy_level = reputation_level

    admission_2025 = _to_int(row[11], default=None) if isinstance(row, list) and len(row) > 11 else None
    max_lottery_2025 = _to_int(row[12], default=None) if isinstance(row, list) and len(row) > 12 else None
    low_rate = _to_float(row[3]) if isinstance(row, list) and len(row) > 3 else None
    high_rate = _to_float(row[4]) if isinstance(row, list) and len(row) > 4 else None
    has_rate = low_rate is not None or high_rate is not None or (admission_2025 is not None and max_lottery_2025 is not None)
    has_enrollment = admission_2025 is not None or max_lottery_2025 is not None
    rate_level = source_level if has_rate else "ai-draft"
    enrollment_level = source_level if has_enrollment else "ai-draft"

    return {
        "enrollment": {"level": enrollment_level, "url": source_url or None, "year": 2025 if has_enrollment else None},
        "rate": {"level": rate_level, "url": source_url or None, "year": 2025 if has_rate else None},
        "reputation": {"level": reputation_level, "url": source_url or None, "year": None},
        "philosophy": {"level": philosophy_level, "url": source_url or None, "year": None},
    }


def _build_honesty_overview(payload: dict, limit: int = 20) -> dict:
    sd = _safe_list(_safe_obj(payload).get("SD"))
    pr = _safe_obj(_safe_obj(payload).get("PR"))
    field_totals = {
        "enrollment": {k: 0 for k in _SOURCE_LEVELS},
        "rate": {k: 0 for k in _SOURCE_LEVELS},
        "reputation": {k: 0 for k in _SOURCE_LEVELS},
        "philosophy": {k: 0 for k in _SOURCE_LEVELS},
    }
    school_rows = []
    for row in sd:
        if not isinstance(row, list) or not row:
            continue
        name = str(row[0] or "").strip()
        if not name:
            continue
        p = _infer_school_provenance(row, pr.get(name))
        for field_key in ("enrollment", "rate", "reputation", "philosophy"):
            lvl = _safe_source_level(_safe_obj(p.get(field_key)).get("level"))
            field_totals[field_key][lvl] += 1
        school_rows.append(
            {
                "name": name,
                "district": str(row[1] or "") if len(row) > 1 else "",
                "type": str(row[2] or "") if len(row) > 2 else "",
                "tier": _normalize_tier(row[10] if len(row) > 10 else "T3"),
                "recommend": _to_int(row[5], default=3) or 3,
                "status": str(row[9] or "normal") if len(row) > 9 else "normal",
                "provenance": p,
            }
        )

    total_schools = len(school_rows) or 1

    def ratio(v):
        return round(v * 1000 / total_schools) / 10

    field_ratios = {}
    for f in ("enrollment", "rate", "reputation", "philosophy"):
        field_ratios[f] = {k: ratio(field_totals[f][k]) for k in _SOURCE_LEVELS}

    priority_rows = []
    for it in school_rows:
        if it["tier"] != "T1":
            continue
        p = it["provenance"]
        rep_level = _safe_obj(p.get("reputation")).get("level")
        rate_level = _safe_obj(p.get("rate")).get("level")
        enrollment_level = _safe_obj(p.get("enrollment")).get("level")
        score = 0
        reasons = []
        if rep_level == "ai-draft":
            score += 4
            reasons.append("口碑为 ai-draft")
        if rate_level == "ai-draft":
            score += 3
            reasons.append("中签率来源缺失")
        if enrollment_level == "ai-draft":
            score += 2
            reasons.append("录取数来源缺失")
        if str(it.get("status")) == "hot":
            score += 1
            reasons.append("热度高（hot）")
        score += max(0, min(3, int(it.get("recommend") or 3) - 3))
        if score <= 0:
            continue
        priority_rows.append(
            {
                "name": it["name"],
                "district": it["district"],
                "type": it["type"],
                "tier": it["tier"],
                "priorityScore": score,
                "reasons": reasons,
                "provenance": p,
            }
        )
    priority_rows.sort(key=lambda x: (-int(x.get("priorityScore") or 0), x.get("district", ""), x.get("name", "")))
    lim = max(1, min(100, int(limit or 20)))
    top_priority = priority_rows[:lim]

    return {
        "totalSchools": len(school_rows),
        "fieldTotals": field_totals,
        "fieldRatios": field_ratios,
        "t1Total": len([x for x in school_rows if x["tier"] == "T1"]),
        "t1NeedReplaceCount": len(priority_rows),
        "t1Priority": top_priority,
    }


def _build_admission_coverage_summary(items: list[dict]) -> dict:
    total = len(items)
    official = len([x for x in items if str(x.get("admission2025Status") or "") == "official"])
    candidate = len([x for x in items if str(x.get("admission2025Status") or "") == "candidate"])
    missing = len([x for x in items if str(x.get("admission2025Status") or "") == "missing"])
    not_applicable = len([x for x in items if str(x.get("admission2025Status") or "") == "not_applicable"])
    return {
        "total": total,
        "official": official,
        "candidate": candidate,
        "missing": missing,
        "notApplicable": not_applicable,
    }


def _build_honesty_school_detail(payload: dict, school_name: str) -> dict:
    sd = _safe_list(_safe_obj(payload).get("SD"))
    pr = _safe_obj(_safe_obj(payload).get("PR"))
    for row in sd:
        if not (isinstance(row, list) and row and str(row[0]) == school_name):
            continue
        p = _infer_school_provenance(row, pr.get(school_name))
        return {
            "name": school_name,
            "district": str(row[1] or "") if len(row) > 1 else "",
            "type": str(row[2] or "") if len(row) > 2 else "",
            "tier": _normalize_tier(row[10] if len(row) > 10 else "T3"),
            "sourceUrl": str(row[13] or "") if len(row) > 13 else "",
            "provenance": p,
            "sourceLevel": _safe_source_level(_safe_obj(pr.get(school_name)).get("sourceLevel")),
            "sourceNote": str(_safe_obj(pr.get(school_name)).get("sourceNote") or ""),
        }
    raise ValueError("学校不存在")


def _extract_json_object(text: str):
    raw = str(text or "").strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        pass
    i = raw.find("{")
    j = raw.rfind("}")
    if i >= 0 and j > i:
        try:
            return json.loads(raw[i : j + 1])
        except Exception:
            return None
    return None


def _normalize_chat_structured(structured: dict, school_hints: list[dict]) -> dict:
    s = _safe_obj(structured)
    assessment = _safe_obj(s.get("assessment"))
    six_layers = _safe_obj(assessment.get("six_layers"))
    confidence = _safe_obj(assessment.get("confidence"))
    score_raw = confidence.get("score")
    try:
        score = float(score_raw)
    except Exception:
        score = 0.5
    score = min(1.0, max(0.0, score))
    out = {
        "next_questions": [str(x) for x in _safe_list(s.get("next_questions")) if str(x).strip()][:1],
        "candidate_schools": _safe_list(s.get("candidate_schools"))[:8],
        "paths": _safe_list(s.get("paths"))[:5],
        "risk_alerts": [str(x) for x in _safe_list(s.get("risk_alerts")) if str(x).strip()][:6],
        "action_items": [str(x) for x in _safe_list(s.get("action_items")) if str(x).strip()][:6],
        "assessment": {
            "one_liner": str(assessment.get("one_liner") or "").strip(),
            "school_portrait": str(assessment.get("school_portrait") or "").strip(),
            "comparison_conclusion": str(assessment.get("comparison_conclusion") or "").strip(),
            "key_evidence": [str(x) for x in _safe_list(assessment.get("key_evidence")) if str(x).strip()][:6],
            "advantages": [str(x) for x in _safe_list(assessment.get("advantages")) if str(x).strip()][:6],
            "risks": [str(x) for x in _safe_list(assessment.get("risks")) if str(x).strip()][:6],
            "family_fit_advice": str(assessment.get("family_fit_advice") or "").strip(),
            "confidence": {
                "level": str(confidence.get("level") or "medium"),
                "score": score,
                "reason": str(confidence.get("reason") or "").strip(),
            },
            "missing_info": [str(x) for x in _safe_list(assessment.get("missing_info")) if str(x).strip()][:8],
            "six_layers": {
                "goal_layer": str(six_layers.get("goal_layer") or "").strip(),
                "mechanism_layer": str(six_layers.get("mechanism_layer") or "").strip(),
                "ai_integration_layer": str(six_layers.get("ai_integration_layer") or "").strip(),
                "cognitive_layer": str(six_layers.get("cognitive_layer") or "").strip(),
                "constraints_layer": str(six_layers.get("constraints_layer") or "").strip(),
                "family_fit_layer": str(six_layers.get("family_fit_layer") or "").strip(),
            },
        },
        "school_reports": _safe_list(s.get("school_reports"))[:6],
        "evidence_chain": _safe_list(s.get("evidence_chain"))[:12],
    }
    if not out["candidate_schools"]:
        out["candidate_schools"] = school_hints[:5]
    return out


def _call_reasoner(model_messages: list[dict], profile: dict, school_hints: list[dict], evidence_pack: list[dict]) -> dict:
    api_key = (os.getenv("DASHSCOPE_API_KEY") or os.getenv("QWEN_API_KEY") or os.getenv("OPENAI_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("LLM 未配置：缺少 DASHSCOPE_API_KEY / QWEN_API_KEY")
    model = (os.getenv("QWEN_MODEL") or os.getenv("DASHSCOPE_MODEL") or os.getenv("OPENAI_MODEL") or "qwen-turbo-latest").strip()
    base_url = (os.getenv("DASHSCOPE_BASE_URL") or "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
    endpoint = f"{base_url}/chat/completions"
    system_prompt = (
        "你是一个面向中国家庭的择校决策助手，请只输出 JSON。\n"
        "核心任务：基于学校信息、家庭约束与孩子画像，帮助用户判断学校真正相信什么、实际上如何运转、是否适配该家庭。\n"
        "行为原则：适配优先；机制优先于口号；先证据后结论；不确定就明确标注；语言克制、结构化。\n"
        "评估框架：教育目标层、学习机制层、AI整合层、认知能力层、现实约束层、家庭适配层。\n"
        "回答顺序强制：先直接回答用户问题，先给明确建议与判断，再补充依据。\n"
        "禁止一上来连续反问；如关键信息缺失，最多提出1个澄清问题，且放在回复末尾。\n"
        "禁止给“最好/第一名”绝对结论；证据不足时必须写“需核验”。\n"
        "JSON 顶层字段：reply, structured。\n"
        "structured 必须包含：next_questions, candidate_schools, paths, risk_alerts, action_items, assessment, school_reports, evidence_chain。\n"
        "assessment 必须包含：one_liner, school_portrait, comparison_conclusion, key_evidence[], advantages[], risks[], family_fit_advice, confidence{level,score,reason}, missing_info[], six_layers{goal_layer,mechanism_layer,ai_integration_layer,cognitive_layer,constraints_layer,family_fit_layer}。\n"
        "school_reports 每项包含：school, advantages[], concerns[], judgement, scores(respect_individuality/homework_pressure_fit/sports_health/creativity_teaching)。\n"
        "evidence_chain 每项包含：school, claim, source_type, title, url, snippet。"
    )
    user_prompt = (
        f"家庭画像：{_profile_brief(profile)}\n"
        f"区域候选学校：{json.dumps(school_hints, ensure_ascii=False)}\n"
        f"可用证据包：{json.dumps(evidence_pack, ensure_ascii=False)}\n"
        "请基于这些信息输出结论。"
    )
    def _invoke_chat(payload_obj: dict) -> tuple[dict, str]:
        req = urllib.request.Request(
            endpoint,
            method="POST",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Authorization": f"Bearer {api_key}",
            },
            data=json.dumps(payload_obj, ensure_ascii=False).encode("utf-8"),
        )
        with urllib.request.urlopen(req, timeout=35) as resp:
            raw_text = resp.read().decode("utf-8", errors="ignore")
        data_obj = json.loads(raw_text or "{}")
        choice0 = _safe_obj((_safe_list(data_obj.get("choices")) or [{}])[0])
        msg_obj = _safe_obj(choice0.get("message"))
        return data_obj, _to_str(msg_obj.get("content"))

    payload = {
        "model": model,
        "response_format": {"type": "json_object"},
        "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}, *model_messages],
        "temperature": 0.35,
        "max_tokens": 900,
    }
    data, content = _invoke_chat(payload)
    parsed = _extract_json_object(content)
    # Some compatible gateways/models may ignore response_format or return plain text.
    # Retry once with stricter last-message constraint and no response_format field.
    if not isinstance(parsed, dict):
        retry_messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
            *model_messages,
            {
                "role": "user",
                "content": (
                    "你上一条未按要求返回JSON。现在请只返回一个 JSON 对象，"
                    "不要代码块，不要解释文字，不要任何前后缀。"
                ),
            },
        ]
        retry_payload = {
            "model": model,
            "messages": retry_messages,
            "temperature": 0.2,
            "max_tokens": 900,
        }
        data2, content2 = _invoke_chat(retry_payload)
        parsed2 = _extract_json_object(content2)
        if isinstance(parsed2, dict):
            parsed = parsed2
            data = data2
            content = content2
    if not isinstance(parsed, dict):
        raw_reply = str(content or "").strip()
        if not raw_reply:
            raw_reply = "我已收到你的问题，但这次模型输出格式异常。请再发一次，我会直接给你结论与建议。"
        return {
            "mode": "reasoner-raw",
            "reply": raw_reply,
            "structured": _normalize_chat_structured({}, school_hints),
            "modelUsed": data.get("model") or model,
        }
    reply = str(parsed.get("reply") or "").strip() or "我已完成初步分析。"
    structured = _normalize_chat_structured(_safe_obj(parsed.get("structured")), school_hints)
    return {"mode": "reasoner", "reply": reply, "structured": structured, "modelUsed": data.get("model") or model}


def _batch_update_school_tiers(updates: list[dict]) -> dict:
    payload = _get_bootstrap_payload()
    sd = payload.get("SD", [])
    update_map = {}
    for item in updates:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        tier = _normalize_tier(item.get("tier"))
        if not name:
            continue
        update_map[name] = tier
    if not update_map:
        return {"updated": 0, "missing": []}

    updated = 0
    existing_names = set()
    for i, row in enumerate(sd):
        if not isinstance(row, list) or not row:
            continue
        name = row[0]
        existing_names.add(name)
        if name in update_map:
            while len(row) <= 10:
                row.append(None)
            old_tier = _normalize_tier(row[10])
            new_tier = update_map[name]
            if old_tier != new_tier:
                row[10] = new_tier
                sd[i] = row
                updated += 1
    missing = [n for n in update_map.keys() if n not in existing_names]
    if updated > 0:
        payload["SD"] = sd
        _replace_payload(payload)
    return {"updated": updated, "missing": missing}


def _apply_school_changes(changes: dict) -> dict:
    payload = _get_bootstrap_payload()
    sd = payload.get("SD", [])
    pr = payload.get("PR", {})
    tf = payload.get("TF", {})

    tier_updates = changes.get("tierUpdates") or []
    rename_updates = changes.get("renameUpdates") or []
    delete_names = changes.get("deleteNames") or []
    if not isinstance(tier_updates, list):
        raise ValueError("tierUpdates 必须是数组")
    if not isinstance(rename_updates, list):
        raise ValueError("renameUpdates 必须是数组")
    if not isinstance(delete_names, list):
        raise ValueError("deleteNames 必须是数组")

    tier_map = {}
    for item in tier_updates:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        tier = _normalize_tier(item.get("tier"))
        if name:
            tier_map[name] = tier

    rename_map = {}
    for item in rename_updates:
        if not isinstance(item, dict):
            continue
        old_name = str(item.get("oldName") or "").strip()
        new_name = str(item.get("newName") or "").strip()
        if old_name and new_name and old_name != new_name:
            rename_map[old_name] = new_name

    delete_set = {str(x).strip() for x in delete_names if str(x).strip()}
    existing_names = {row[0] for row in sd if isinstance(row, list) and row}
    rename_conflicts = []
    for old_name, new_name in rename_map.items():
        # 仅允许改成不存在的名字，避免覆盖另一所学校
        if new_name in existing_names and new_name != old_name:
            rename_conflicts.append({"oldName": old_name, "newName": new_name, "reason": "newName 已存在"})
    if rename_conflicts:
        raise ValueError("存在重名冲突，请先处理： " + ", ".join([f"{x['oldName']}->{x['newName']}" for x in rename_conflicts]))

    new_sd = []
    tier_updated = 0
    renamed = 0
    deleted = 0
    for row in sd:
        if not isinstance(row, list) or not row:
            continue
        name = row[0]
        if name in delete_set:
            deleted += 1
            continue
        if name in tier_map:
            while len(row) <= 10:
                row.append(None)
            old_tier = _normalize_tier(row[10])
            new_tier = tier_map[name]
            if old_tier != new_tier:
                row[10] = new_tier
                # 同步 PR.tag 中的梯队文本，避免“列表是T1但详情口碑标签还是T2”
                if isinstance(pr, dict):
                    p_node = pr.get(name)
                    if isinstance(p_node, dict):
                        old_tag = p_node.get("tag")
                        if isinstance(old_tag, str) and old_tag.strip():
                            if re.search(r"\bT[123]\b", old_tag):
                                p_node["tag"] = re.sub(r"\bT[123]\b", new_tier, old_tag)
                            else:
                                p_node["tag"] = (old_tag + "·" + new_tier).strip("·")
                            pr[name] = p_node
                tier_updated += 1
        if name in rename_map:
            old_name = name
            new_name = rename_map[name]
            row[0] = new_name
            if isinstance(pr, dict) and old_name in pr:
                pr[new_name] = pr.pop(old_name)
            if isinstance(tf, dict) and old_name in tf:
                tf[new_name] = tf.pop(old_name)
            renamed += 1
        new_sd.append(row)

    for name in delete_set:
        if isinstance(pr, dict):
            pr.pop(name, None)
        if isinstance(tf, dict):
            tf.pop(name, None)

    missing_tier_names = [n for n in tier_map.keys() if n not in existing_names]
    missing_rename_names = [n for n in rename_map.keys() if n not in existing_names]
    missing_delete_names = [n for n in delete_set if n not in existing_names]

    changed = renamed > 0 or deleted > 0 or tier_updated > 0 or (len(delete_set) > len(missing_delete_names))
    if changed:
        payload["SD"] = new_sd
        payload["PR"] = pr
        payload["TF"] = tf
        _replace_payload(payload)

    return {
        "renamed": renamed,
        "tierUpdated": tier_updated,
        "deleted": deleted,
        "missingRenameNames": missing_rename_names,
        "missingTierNames": missing_tier_names,
        "missingDeleteNames": missing_delete_names,
    }


def _extract_school_from_xlsx(content: bytes) -> list[list]:
    if load_workbook is None:
        raise ValueError("openpyxl 未安装，无法解析 .xlsx")
    wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = None
    for r in rows:
        vals = [str(x).strip() if x is not None else "" for x in r]
        if any(vals):
            header = vals
            break
    if not header:
        return []

    aliases = {
        "name": {"name", "学校名", "学校名称"},
        "district": {"district", "区", "区域"},
        "type": {"type", "学校类型", "类型", "公办民办"},
        "tier": {"tier", "梯队"},
        "lottery_low": {"lotterylow", "中签率低", "中签率下限", "low"},
        "lottery_high": {"lotteryhigh", "中签率高", "中签率上限", "high"},
        "recommend": {"recommend", "推荐度"},
        "desc": {"desc", "简介", "描述"},
        "lat": {"lat", "纬度"},
        "lng": {"lng", "经度"},
        "status": {"status", "状态"},
        "admission_2025": {"admission2025", "2025录取数", "录取数2025"},
        "max_lottery_2025": {"maxlottery2025", "2025最大摇号数", "最大摇号数2025"},
        "source_url": {"sourceurl", "参考信息来源网址", "来源网址", "source"},
    }

    def norm_key(s: str) -> str:
        return re.sub(r"\s+", "", s).lower().replace("-", "").replace("_", "")

    idx = {}
    norm_headers = [norm_key(h) for h in header]
    for canon, keys in aliases.items():
        keyset = {norm_key(k) for k in keys}
        for i, h in enumerate(norm_headers):
            if h in keyset:
                idx[canon] = i
                break

    if "name" not in idx:
        raise ValueError("Excel 缺少必需列：学校名称（name）")

    result = []
    for r in rows:
        vals = list(r)
        if not any(v is not None and str(v).strip() != "" for v in vals):
            continue

        def get(canon, default=None):
            i = idx.get(canon)
            if i is None or i >= len(vals):
                return default
            return vals[i]

        name = str(get("name", "")).strip()
        if not name:
            continue
        district = str(get("district", "")).strip()
        s_type = _normalize_type(get("type", "pub"))
        tier = _normalize_tier(get("tier", "T3"))
        low = _to_float(get("lottery_low"))
        high = _to_float(get("lottery_high"))
        recommend = _to_int(get("recommend"), default=3)
        desc = str(get("desc", "") or "").strip()
        lat = _to_float(get("lat"))
        lng = _to_float(get("lng"))
        status = _normalize_status(get("status", "normal"))
        admission_2025 = _to_int(get("admission_2025"), default=None)
        max_lottery_2025 = _to_int(get("max_lottery_2025"), default=None)
        source_url = str(get("source_url", "") or "").strip()
        row = [
            name,
            district,
            s_type,
            low,
            high,
            recommend,
            desc,
            lat,
            lng,
            status,
            tier,
            admission_2025,
            max_lottery_2025,
            source_url,
        ]
        result.append(row)
    return result


def _insert_proposals(source: str, evidence_url: str, proposals: list[dict]) -> dict:
    created = 0
    skipped = 0
    with _db_conn() as conn:
        for p in proposals:
            p_type = str(p.get("proposalType", "")).strip()
            p_key = str(p.get("proposalKey", "")).strip()
            if not p_type or not p_key or "newValue" not in p:
                skipped += 1
                continue
            try:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO proposals(
                      source, proposal_type, proposal_key, new_value_json, evidence_url, status, created_at
                    ) VALUES(?, ?, ?, ?, ?, 'pending', ?)
                    """,
                    (
                        source,
                        p_type,
                        p_key,
                        json.dumps(p["newValue"], ensure_ascii=False),
                        p.get("evidenceUrl") or evidence_url,
                        _now(),
                    ),
                )
                if conn.total_changes > created:
                    created += 1
                else:
                    skipped += 1
            except sqlite3.IntegrityError:
                skipped += 1
        conn.commit()
    return {"created": created, "skipped": skipped}


def _fetch_text(url: str) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": "PrimarySchoolAdvisorBot/1.0"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        return resp.read().decode("utf-8", errors="ignore")


def _to_count(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().lower().replace(",", "")
    if not s:
        return None
    m = re.match(r"^(\d+(?:\.\d+)?)\s*([wk万千k]?)$", s)
    if not m:
        try:
            return int(float(s))
        except ValueError:
            return None
    num = float(m.group(1))
    unit = m.group(2)
    if unit in ("w", "万"):
        num *= 10000
    elif unit in ("k", "千"):
        num *= 1000
    return int(num)


def _extract_xhs_meta(url: str) -> dict:
    html = _fetch_text(url)
    title = ""
    m = re.search(r'<meta[^>]+property=["\']og:title["\'][^>]*content=["\'](.*?)["\']', html, flags=re.I | re.S)
    if m:
        title = re.sub(r"\s+", " ", m.group(1)).strip()
    if not title:
        m = re.search(r"<title[^>]*>(.*?)</title>", html, flags=re.I | re.S)
        if m:
            title = re.sub(r"\s+", " ", m.group(1)).strip()

    published_at = ""
    date_patterns = [
        r"(20\d{2}-\d{1,2}-\d{1,2})",
        r"(20\d{2}/\d{1,2}/\d{1,2})",
        r"(20\d{2}年\d{1,2}月\d{1,2}日)",
    ]
    for p in date_patterns:
        mm = re.search(p, html)
        if mm:
            published_at = mm.group(1)
            break

    def pick_int(patterns):
        for p in patterns:
            mm = re.search(p, html, flags=re.I)
            if mm:
                val = _to_count(mm.group(1))
                if val is not None:
                    return val
        return None

    like_count = pick_int(
        [
            r'"liked_count"\s*:\s*"?([0-9]+(?:\.[0-9]+)?[wk万千k]?)"?',
            r'"like_count"\s*:\s*"?([0-9]+(?:\.[0-9]+)?[wk万千k]?)"?',
            r"点赞[^0-9]{0,6}([0-9]+(?:\.[0-9]+)?[wk万千k]?)",
        ]
    )
    fav_count = pick_int(
        [
            r'"collected_count"\s*:\s*"?([0-9]+(?:\.[0-9]+)?[wk万千k]?)"?',
            r'"collect_count"\s*:\s*"?([0-9]+(?:\.[0-9]+)?[wk万千k]?)"?',
            r"收藏[^0-9]{0,6}([0-9]+(?:\.[0-9]+)?[wk万千k]?)",
        ]
    )

    return {"title": title, "publishedAt": published_at, "likeCount": like_count, "favoriteCount": fav_count}


def _collect_xhs_proposals(items: list[dict], purpose: str) -> dict:
    proposals = []
    errors = []
    for i, it in enumerate(items):
        if not isinstance(it, dict):
            errors.append({"index": i, "error": "item 必须是对象"})
            continue
        school_name = str(it.get("schoolName") or "").strip()
        url = str(it.get("url") or "").strip()
        if not school_name or not url:
            errors.append({"index": i, "error": "schoolName/url 不能为空"})
            continue
        tier = str(it.get("tier") or "").strip().upper()
        if tier:
            tier = _normalize_tier(tier)
        admission_2025 = _to_int(it.get("admission2025"), default=None)
        max_lottery_2025 = _to_int(it.get("maxLottery2025"), default=None)
        note = str(it.get("note") or "").strip()

        meta = {}
        fetch_error = ""
        try:
            meta = _extract_xhs_meta(url)
        except Exception as e:
            fetch_error = str(e)
            meta = {"title": "", "publishedAt": "", "likeCount": None, "favoriteCount": None}

        new_value = {
            "tier": tier or None,
            "admission2025": admission_2025,
            "maxLottery2025": max_lottery_2025,
            "sourceUrl": url,
            "xhsMeta": {
                **meta,
                "url": url,
                "purpose": purpose,
                "note": note,
                "fetchError": fetch_error,
                "capturedAt": _now(),
            },
        }
        proposals.append(
            {
                "proposalType": "patch_school_fields",
                "proposalKey": school_name,
                "newValue": new_value,
                "evidenceUrl": url,
            }
        )
    result = _insert_proposals(f"xhs:{purpose}", "", proposals)
    return {**result, "errors": errors, "submitted": len(items)}


def _collect_policy_html(source: dict) -> list[dict]:
    html = _fetch_text(source["url"])
    title_match = re.search(r"<title[^>]*>(.*?)</title>", html, flags=re.I | re.S)
    page_title = re.sub(r"\s+", " ", title_match.group(1).strip()) if title_match else source["name"]
    keywords = source.get("keywords") or []
    if keywords and not any(k in html for k in keywords):
        return []
    dates = set(re.findall(r"\b(20\d{2}-\d{1,2}-\d{1,2})\b", html))
    zh_dates = re.findall(r"(20\d{2}年\d{1,2}月\d{1,2}日)", html)
    dates.update(zh_dates)
    proposals = []
    for i, d in enumerate(sorted(dates)[:30]):
        key = f"{source['id']}::{d}::{i}"
        proposals.append(
            {
                "proposalType": "add_policy_event",
                "proposalKey": key,
                "newValue": {"date": d, "title": page_title, "source": source["name"], "url": source["url"]},
                "evidenceUrl": source["url"],
            }
        )
    return proposals


def _collect_school_json(source: dict) -> list[dict]:
    raw = _fetch_text(source["url"])
    data = json.loads(raw)
    if not isinstance(data, list):
        return []
    m = source.get("mapping") or {}
    proposals = []
    for item in data:
        if not isinstance(item, dict):
            continue
        name = item.get(m.get("name", "name"))
        if not name:
            continue
        row = [
            name,
            item.get(m.get("district", "district"), ""),
            item.get(m.get("schoolType", "type"), "pub"),
            item.get(m.get("lotteryLow", "lotteryLow")),
            item.get(m.get("lotteryHigh", "lotteryHigh")),
            item.get(m.get("recommend", "recommend"), 3),
            item.get(m.get("desc", "desc"), ""),
            item.get(m.get("lat", "lat")),
            item.get(m.get("lng", "lng")),
            item.get(m.get("status", "status"), "normal"),
            item.get(m.get("tier", "tier"), "T3"),
        ]
        proposals.append(
            {
                "proposalType": "upsert_sd",
                "proposalKey": str(name),
                "newValue": row,
                "evidenceUrl": source["url"],
            }
        )
    return proposals


def _run_collect_once() -> dict:
    sources = _load_sources()
    summary = {"scannedSources": 0, "created": 0, "skipped": 0, "errors": []}
    for s in sources:
        if not s.get("enabled"):
            continue
        summary["scannedSources"] += 1
        try:
            proposals = []
            if s.get("type") == "policy_html":
                proposals = _collect_policy_html(s)
            elif s.get("type") == "school_json":
                proposals = _collect_school_json(s)
            result = _insert_proposals(s.get("name") or s.get("id") or "collector", s.get("url", ""), proposals)
            summary["created"] += result["created"]
            summary["skipped"] += result["skipped"]
        except Exception as e:
            summary["errors"].append({"source": s.get("id"), "error": str(e)})
    return summary


def _apply_single_proposal(payload: dict, row: sqlite3.Row) -> bool:
    changed = False
    p_type = row["proposal_type"]
    p_key = row["proposal_key"]
    new_value = json.loads(row["new_value_json"])

    if p_type == "upsert_sd":
        sd = payload["SD"]
        if not isinstance(new_value, list) or not new_value:
            return False
        for i, school in enumerate(sd):
            if isinstance(school, list) and school and school[0] == p_key:
                sd[i] = new_value
                return True
        sd.append(new_value)
        return True
    if p_type == "patch_school_fields":
        if not isinstance(new_value, dict):
            return False
        sd = payload["SD"]
        for i, school in enumerate(sd):
            if not (isinstance(school, list) and school and school[0] == p_key):
                continue
            while len(school) <= 13:
                school.append(None)
            changed = False
            tier_val = new_value.get("tier")
            if isinstance(tier_val, str) and tier_val.strip():
                t = _normalize_tier(tier_val)
                if school[10] != t:
                    school[10] = t
                    changed = True
            admission = _to_int(new_value.get("admission2025"), default=None)
            if admission is not None and school[11] != admission:
                school[11] = admission
                changed = True
            max_lottery = _to_int(new_value.get("maxLottery2025"), default=None)
            if max_lottery is not None and school[12] != max_lottery:
                school[12] = max_lottery
                changed = True

            source_url = str(new_value.get("sourceUrl") or "").strip()
            old_source = str(school[13] or "").strip()
            if source_url:
                if not old_source:
                    school[13] = source_url
                    changed = True
                elif source_url not in old_source:
                    school[13] = old_source + "；" + source_url
                    changed = True

            xhs_meta = new_value.get("xhsMeta")
            if isinstance(xhs_meta, dict):
                pr = payload.get("PR")
                if isinstance(pr, dict):
                    node = pr.get(p_key)
                    if not isinstance(node, dict):
                        node = {}
                    old = node.get("xhsSignals")
                    if not isinstance(old, list):
                        old = []
                    old.append(xhs_meta)
                    node["xhsSignals"] = old[-20:]
                    pr[p_key] = node
                    payload["PR"] = pr
                    changed = True
            if changed:
                sd[i] = school
            return changed
        return False
    if p_type == "upsert_pr":
        payload["PR"][p_key] = new_value
        return True
    if p_type == "patch_pr_fields":
        if not isinstance(new_value, dict):
            return False
        pr = payload.get("PR")
        if not isinstance(pr, dict):
            pr = {}
        node = pr.get(p_key)
        if not isinstance(node, dict):
            node = {}
        changed = False
        for k, v in new_value.items():
            if node.get(k) != v:
                node[k] = v
                changed = True
        if changed:
            pr[p_key] = node
            payload["PR"] = pr
        return changed
    if p_type == "upsert_tf":
        payload["TF"][p_key] = new_value
        return True
    if p_type == "patch_tf_fields":
        if not isinstance(new_value, dict):
            return False
        tf = payload.get("TF")
        if not isinstance(tf, dict):
            tf = {}
        node = tf.get(p_key)
        if not isinstance(node, dict):
            node = {}
        changed = False
        for k, v in new_value.items():
            if node.get(k) != v:
                node[k] = v
                changed = True
        if changed:
            tf[p_key] = node
            payload["TF"] = tf
        return changed
    if p_type == "set_dn":
        payload["DN"][p_key] = new_value
        return True
    if p_type == "add_policy_event":
        with _db_conn() as conn:
            conn.execute(
                """
                INSERT OR IGNORE INTO policy_events(
                  event_key, event_date, title, source, evidence_url, payload_json, created_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    p_key,
                    str(new_value.get("date") or ""),
                    str(new_value.get("title") or ""),
                    str(new_value.get("source") or row["source"]),
                    row["evidence_url"],
                    json.dumps(new_value, ensure_ascii=False),
                    _now(),
                ),
            )
            conn.commit()
        return False
    return changed


def _read_json_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length > 0 else b"{}"
    return json.loads(raw.decode("utf-8"))


def _serve_admin_static(handler: BaseHTTPRequestHandler, p: str) -> bool:
    if p == "/admin" or p == "/admin/":
        p = "/admin/index.html"
    if not p.startswith("/admin/"):
        return False
    rel = p[len("/admin/") :]
    file_path = (ADMIN_DIR / rel).resolve()
    if not str(file_path).startswith(str(ADMIN_DIR.resolve())) or not file_path.exists():
        _json_response(handler, 404, {"error": "Not Found"})
        return True
    if file_path.suffix == ".html":
        ctype = "text/html; charset=utf-8"
    elif file_path.suffix == ".js":
        ctype = "application/javascript; charset=utf-8"
    elif file_path.suffix == ".css":
        ctype = "text/css; charset=utf-8"
    else:
        ctype = "text/plain; charset=utf-8"
    _text_response(handler, 200, file_path.read_text(encoding="utf-8"), ctype)
    return True


class Handler(BaseHTTPRequestHandler):
    def _redirect(self, location: str, code: int = 302) -> None:
        self.send_response(code)
        self.send_header("Location", location)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self) -> None:
        p = urlparse(self.path)
        if p.path in ("/", "/index.html"):
            self._redirect("/admin/")
            return
        if _serve_admin_static(self, p.path):
            return
        if p.path == "/api/health":
            _json_response(self, 200, {"ok": True, "dbPath": str(DB_PATH), "sourcesPath": str(SOURCES_PATH)})
            return
        if p.path == "/api/bootstrap":
            _json_response(self, 200, _get_bootstrap_payload())
            return
        if p.path == "/api/schools":
            payload = _get_bootstrap_payload()
            q = parse_qs(p.query)
            district = q.get("district", [""])[0]
            school_type = q.get("type", [""])[0]
            schools = payload["SD"]
            if district:
                schools = [s for s in schools if len(s) > 1 and s[1] == district]
            if school_type:
                schools = [s for s in schools if len(s) > 2 and s[2] == school_type]
            _json_response(self, 200, {"count": len(schools), "items": schools})
            return
        if p.path == "/api/policy-events":
            with _db_conn() as conn:
                rows = conn.execute(
                    """
                    SELECT id, event_key, event_date, title, source, evidence_url, payload_json, created_at
                    FROM policy_events
                    ORDER BY id DESC
                    LIMIT 300
                    """
                ).fetchall()
            items = []
            for r in rows:
                items.append(
                    {
                        "id": r["id"],
                        "eventKey": r["event_key"],
                        "date": r["event_date"],
                        "title": r["title"],
                        "source": r["source"],
                        "evidenceUrl": r["evidence_url"],
                        "payload": json.loads(r["payload_json"]),
                        "createdAt": r["created_at"],
                    }
                )
            _json_response(self, 200, {"count": len(items), "items": items})
            return
        if p.path == "/api/admin/sources":
            _json_response(self, 200, {"sources": _load_sources(), "path": str(SOURCES_PATH)})
            return
        if p.path == "/api/admin/proposals":
            q = parse_qs(p.query)
            status = q.get("status", ["pending"])[0]
            with _db_conn() as conn:
                rows = conn.execute(
                    """
                    SELECT id, source, proposal_type, proposal_key, new_value_json, evidence_url, status, note, created_at, reviewed_at
                    FROM proposals
                    WHERE (? = 'all' OR status = ?)
                    ORDER BY id DESC
                    LIMIT 1000
                    """,
                    (status, status),
                ).fetchall()
            items = []
            for r in rows:
                items.append(
                    {
                        "id": r["id"],
                        "source": r["source"],
                        "proposalType": r["proposal_type"],
                        "proposalKey": r["proposal_key"],
                        "newValue": json.loads(r["new_value_json"]),
                        "evidenceUrl": r["evidence_url"],
                        "status": r["status"],
                        "note": r["note"],
                        "createdAt": r["created_at"],
                        "reviewedAt": r["reviewed_at"],
                    }
                )
            _json_response(self, 200, {"count": len(items), "items": items})
            return
        if p.path == "/api/admin/schools":
            payload = _get_bootstrap_payload()
            items = _build_school_items(payload)
            coverage = _build_admission_coverage_summary(items)
            q = parse_qs(p.query)
            district = str(q.get("district", [""])[0]).strip()
            school_type = str(q.get("type", [""])[0]).strip()
            tier = str(q.get("tier", [""])[0]).strip().upper()
            keyword = str(q.get("q", [""])[0]).strip().lower()

            def ok(it):
                if district and it["district"] != district:
                    return False
                if school_type and it["type"] != school_type:
                    return False
                if tier and str(it["tier"]).upper() != tier:
                    return False
                if keyword and keyword not in str(it["name"]).lower():
                    return False
                return True

            items = [x for x in items if ok(x)]
            items.sort(key=lambda x: (x.get("district", ""), x.get("tier", ""), x.get("name", "")))
            _json_response(self, 200, {"count": len(items), "items": items, "admissionCoverage": coverage})
            return
        if p.path == "/api/admin/school-detail":
            q = parse_qs(p.query)
            name = str(q.get("name", [""])[0]).strip()
            if not name:
                _json_response(self, 400, {"error": "name 不能为空"})
                return
            payload = _get_bootstrap_payload()
            for it in _build_school_items(payload):
                if it["name"] == name:
                    _json_response(
                        self,
                        200,
                        {
                            "ok": True,
                            "item": it,
                            "fieldSections": _school_field_sections(it),
                            "sources": _collect_school_sources(it),
                        },
                    )
                    return
            _json_response(self, 404, {"error": "学校不存在"})
            return
        if p.path == "/api/admin/school-evidence":
            q = parse_qs(p.query)
            name = str(q.get("name", [""])[0]).strip()
            if not name:
                _json_response(self, 400, {"error": "name 不能为空"})
                return
            limit = _to_int(q.get("limit", ["50"])[0], default=50)
            items = _list_school_evidence(name, limit=limit or 50)
            _json_response(self, 200, {"ok": True, "count": len(items), "items": items})
            return
        if p.path == "/api/admin/search-extract-demo":
            q = parse_qs(p.query)
            name = str(q.get("name", [""])[0]).strip()
            if not name:
                _json_response(self, 400, {"error": "name 不能为空"})
                return
            query_text = str(q.get("query", [""])[0]).strip() or _primary_stage_query(name, "2025", "幼升小", "中签率")
            title = str(q.get("title", [""])[0]).strip()
            snippet = str(q.get("snippet", [""])[0]).strip()
            url = str(q.get("url", [""])[0]).strip()
            result = _extract_candidate_facts_from_search_text(name, query_text, title, snippet, url)
            _json_response(self, 200, {"ok": True, "item": result})
            return
        if p.path == "/api/admin/honesty/overview":
            q = parse_qs(p.query)
            limit = _to_int(q.get("limit", ["20"])[0], default=20) or 20
            payload = _get_bootstrap_payload()
            summary = _build_honesty_overview(payload, limit=limit)
            _json_response(self, 200, {"ok": True, **summary})
            return
        if p.path == "/api/admin/honesty/school-detail":
            q = parse_qs(p.query)
            name = str(q.get("name", [""])[0]).strip()
            if not name:
                _json_response(self, 400, {"error": "name 不能为空"})
                return
            payload = _get_bootstrap_payload()
            try:
                detail = _build_honesty_school_detail(payload, name)
            except ValueError as e:
                _json_response(self, 404, {"error": str(e)})
                return
            _json_response(self, 200, {"ok": True, "item": detail})
            return
        if p.path == "/api/admin/publish-online/status":
            q = parse_qs(p.query)
            task_id = str(q.get("taskId", [""])[0]).strip()
            if not task_id:
                _json_response(self, 400, {"error": "taskId 不能为空"})
                return
            task = _snapshot_publish_task(task_id)
            if not task:
                _json_response(self, 404, {"error": "任务不存在"})
                return
            _json_response(self, 200, {"ok": True, "task": task})
            return
        _json_response(self, 404, {"error": "Not Found"})

    def do_PUT(self) -> None:
        p = urlparse(self.path)
        try:
            body = _read_json_body(self)
        except json.JSONDecodeError:
            _json_response(self, 400, {"error": "请求体不是合法 JSON"})
            return

        if p.path == "/api/bootstrap":
            try:
                _replace_payload(body)
            except ValueError as e:
                _json_response(self, 400, {"error": str(e)})
                return
            _json_response(self, 200, {"ok": True})
            return

        if p.path == "/api/admin/sources":
            sources = body.get("sources")
            if not isinstance(sources, list):
                _json_response(self, 400, {"error": "sources 必须是数组"})
                return
            _save_sources(sources)
            _json_response(self, 200, {"ok": True, "count": len(sources)})
            return
        _json_response(self, 404, {"error": "Not Found"})

    def do_POST(self) -> None:
        p = urlparse(self.path)
        try:
            body = _read_json_body(self)
        except json.JSONDecodeError:
            _json_response(self, 400, {"error": "请求体不是合法 JSON"})
            return

        if p.path == "/api/admin/proposals/import":
            source = str(body.get("source") or "manual")
            evidence_url = str(body.get("evidenceUrl") or "")
            proposals = body.get("proposals")
            if not isinstance(proposals, list):
                _json_response(self, 400, {"error": "proposals 必须是数组"})
                return
            result = _insert_proposals(source, evidence_url, proposals)
            _json_response(self, 200, {"ok": True, **result})
            return

        if p.path == "/api/admin/schools/import-xlsx":
            filename = str(body.get("filename") or "upload.xlsx")
            content_b64 = body.get("contentBase64")
            mode = str(body.get("mode") or "proposals").strip().lower()
            if not isinstance(content_b64, str) or not content_b64:
                _json_response(self, 400, {"error": "contentBase64 不能为空"})
                return
            try:
                content = base64.b64decode(content_b64)
                rows = _extract_school_from_xlsx(content)
            except Exception as e:
                _json_response(self, 400, {"error": f"Excel 解析失败: {e}"})
                return

            if mode == "apply":
                payload = _get_bootstrap_payload()
                changed = 0
                for row in rows:
                    name = row[0]
                    updated = False
                    for i, old in enumerate(payload["SD"]):
                        if isinstance(old, list) and old and old[0] == name:
                            payload["SD"][i] = row
                            updated = True
                            break
                    if not updated:
                        payload["SD"].append(row)
                    changed += 1
                _replace_payload(payload)
                _json_response(self, 200, {"ok": True, "mode": "apply", "updated": changed})
                return

            proposals = []
            for row in rows:
                proposals.append(
                    {
                        "proposalType": "upsert_sd",
                        "proposalKey": row[0],
                        "newValue": row,
                        "evidenceUrl": filename,
                    }
                )
            result = _insert_proposals(f"excel:{filename}", filename, proposals)
            _json_response(self, 200, {"ok": True, "mode": "proposals", **result})
            return

        if p.path == "/api/admin/schools/tier-batch-update":
            updates = body.get("updates")
            if not isinstance(updates, list):
                _json_response(self, 400, {"error": "updates 必须是数组"})
                return
            result = _batch_update_school_tiers(updates)
            _json_response(self, 200, {"ok": True, **result})
            return

        if p.path == "/api/admin/schools/push-changes":
            try:
                result = _apply_school_changes(body)
            except ValueError as e:
                _json_response(self, 400, {"error": str(e)})
                return
            _json_response(self, 200, {"ok": True, **result})
            return

        if p.path == "/api/admin/publish-online":
            message = str(body.get("message") or "").strip()
            if not message:
                message = "chore: publish latest school data"
            mode = str(body.get("mode") or "api").strip().lower() or "api"
            if mode not in ("api", "git"):
                _json_response(self, 400, {"error": "mode 必须是 api 或 git"})
                return
            remote = str(body.get("remote") or "cf").strip() or "cf"
            refspec = str(body.get("refspec") or "cf-main:main").strip() or "cf-main:main"
            git_ssh_command = str(body.get("gitSshCommand") or "").strip()
            cloudflare_url = str(body.get("cloudflareUrl") or "https://school-advisor.pages.dev/api/admin/bootstrap").strip()
            publish_token = str(body.get("publishToken") or os.getenv("CF_PUBLISH_TOKEN", "")).strip()
            if mode == "api" and not publish_token:
                _json_response(self, 400, {"error": "缺少 publishToken（可在前端输入或设置 CF_PUBLISH_TOKEN）"})
                return
            task = _new_publish_task(message, remote, refspec, mode=mode, cloudflare_url=cloudflare_url)
            t = threading.Thread(
                target=_run_publish_task,
                args=(task["taskId"], message, remote, refspec, git_ssh_command, mode, cloudflare_url, publish_token),
                daemon=True,
            )
            t.start()
            _json_response(self, 200, {"ok": True, "taskId": task["taskId"], "task": task})
            return

        if p.path == "/api/admin/xhs/collect-proposals":
            items = body.get("items")
            purpose = str(body.get("purpose") or "tier2026").strip() or "tier2026"
            if not isinstance(items, list) or not items:
                _json_response(self, 400, {"error": "items 必须是非空数组"})
                return
            result = _collect_xhs_proposals(items, purpose)
            _json_response(self, 200, {"ok": True, **result})
            return

        if p.path == "/api/admin/school-evidence/collect":
            school_name = str(body.get("schoolName") or "").strip()
            school_names = body.get("schoolNames")
            top = _to_int(body.get("top"), default=3) or 3
            extra_queries = body.get("queries")
            if school_name:
                targets = [school_name]
            elif isinstance(school_names, list):
                targets = [str(x).strip() for x in school_names if str(x).strip()]
            else:
                targets = []
            if not targets:
                _json_response(self, 400, {"error": "schoolName 或 schoolNames 至少提供一个"})
                return
            if len(targets) > 20:
                _json_response(self, 400, {"error": "单次最多抓取20所学校"})
                return
            out = []
            for n in targets:
                out.append(_collect_school_evidence(n, extra_queries if isinstance(extra_queries, list) else None, top=top))
            _json_response(self, 200, {"ok": True, "count": len(out), "results": out})
            return

        if p.path == "/api/chat/decision":
            if not ENABLE_1V1_DECISION:
                _json_response(
                    self,
                    410,
                    {"ok": False, "code": "feature_disabled", "reply": "1v1 决策顾问暂时下线，功能迭代完成后会重新开放。"},
                )
                return
            messages = _trim_chat_messages(body.get("messages"))
            if not messages:
                _json_response(self, 400, {"error": "messages 不能为空"})
                return
            profile = _safe_obj(body.get("profile"))
            payload = _get_bootstrap_payload()
            school_hints = _district_school_hints(payload, str(profile.get("district") or ""))
            evidence_pack = []
            for s in school_hints[:3]:
                sn = str(s.get("name") or "").strip()
                if not sn:
                    continue
                evidence_pack.append({"schoolName": sn, "evidence": _list_school_evidence(sn, limit=8)})
            try:
                out = _call_reasoner(messages, profile, school_hints, evidence_pack)
                _json_response(self, 200, {"ok": True, **out})
            except Exception as e:
                _json_response(
                    self,
                    503,
                    {
                        "ok": False,
                        "code": "llm_unavailable",
                        "reply": "当前模型连接暂时不可用，请重试一次；恢复后会直接给你完整建议。",
                        "error": str(e),
                    },
                )
            return

        if p.path == "/api/admin/collect/run":
            summary = _run_collect_once()
            _json_response(self, 200, {"ok": True, **summary})
            return

        if p.path == "/api/admin/proposals/review":
            ids = body.get("ids")
            action = str(body.get("action") or "").strip()
            note = str(body.get("note") or "")
            if not isinstance(ids, list) or not ids:
                _json_response(self, 400, {"error": "ids 必须是非空数组"})
                return
            if action not in ("approve", "reject"):
                _json_response(self, 400, {"error": "action 必须是 approve 或 reject"})
                return
            with _db_conn() as conn:
                marks = ",".join("?" for _ in ids)
                rows = conn.execute(
                    f"""
                    SELECT id, source, proposal_type, proposal_key, new_value_json, evidence_url
                    FROM proposals
                    WHERE status='pending' AND id IN ({marks})
                    """,
                    ids,
                ).fetchall()
                if action == "reject":
                    conn.execute(
                        f"""
                        UPDATE proposals
                        SET status='rejected', note=?, reviewed_at=?
                        WHERE status='pending' AND id IN ({marks})
                        """,
                        [note, _now(), *ids],
                    )
                    conn.commit()
                    _json_response(self, 200, {"ok": True, "updated": len(rows), "action": action})
                    return

            payload = _get_bootstrap_payload()
            payload_changed = False
            for r in rows:
                payload_changed = _apply_single_proposal(payload, r) or payload_changed
            if payload_changed:
                _replace_payload(payload)
            with _db_conn() as conn:
                marks = ",".join("?" for _ in ids)
                conn.execute(
                    f"""
                    UPDATE proposals
                    SET status='approved', note=?, reviewed_at=?
                    WHERE status='pending' AND id IN ({marks})
                    """,
                    [note, _now(), *ids],
                )
                conn.commit()
            _json_response(
                self,
                200,
                {"ok": True, "updated": len(rows), "action": action, "payloadChanged": payload_changed},
            )
            return
        _json_response(self, 404, {"error": "Not Found"})

    def log_message(self, fmt: str, *args) -> None:
        return


def main() -> None:
    _init_db()
    _ensure_default_sources()
    parser = argparse.ArgumentParser(description="Primary School Advisor API server")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8787)
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"API running at http://{args.host}:{args.port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
